import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Color, PatternFill
from openpyxl.utils import get_column_letter


class Relatorio_xls:

    def __init__(self, caminho:str, dados):
        self.__caminho = caminho
        self.__dados = dados
        self.__workbook = Workbook()

        self.plan_ativa = self.__workbook.active

        self.preencherPlanilha()

    def setTitulo(self, titulo:str="Relatório") -> None:
        self.plan_ativa.title = titulo

    def salvar(self) -> None:
        self.__workbook.save(self.__caminho)

    def cabecalho(self) -> None:
        tabela = {
            "A2": ["DECRETO", Color(rgb='a4c2f4'), 1],
            "B2": ["DATA DECRETO", Color(rgb='a4c2f4'), 2],
            "C2": ["ORGÃO ORIGEM", Color(rgb='a4c2f4'), 3],
            "D2": ["PROGRAMA ORIGEM", Color(rgb='a4c2f4'), 4],
            "E2": ["AÇÃO ORIGEM", Color(rgb='a4c2f4'), 5],
            "F2": ["PRODUTO ORIGEM", Color(rgb='a4c2f4'), 6],
            "G2": ["META FÍSICA ORIGEM", Color(rgb='a4c2f4'), 7],
            "H2": ["NOVA META FÍSICA ORIGEM", Color(rgb='a4c2f4'), 8],
            "I2": ["ORGÃO DESTINO", Color(rgb='a4c2f4'), 9],
            "J2": ["PROGRAMA DESTINO", Color(rgb='a4c2f4'), 10],
            "K2": ["AÇÃO DESTINO", Color(rgb='a4c2f4'), 11],
            "L2": ["PRODUTO DESTINO", Color(rgb='a4c2f4'), 12],
            "M2": ["META FÍSICA DESTINO", Color(rgb='a4c2f4'), 13],
            "N2": ["NOVA META FÍSICA DESTINO", Color(rgb='a4c2f4'), 14],
            "O2": ["DATA EMAIL INICIAL", Color(rgb='a4c2f4'), 15],
            "P2": ["EMAIL", Color(rgb='a4c2f4'), 16],
            "Q2": ["VALOR REMANEJADO", Color(rgb='a4c2f4'), 17],
        }

        for cel, value in tabela.items():
            self.plan_ativa[cel].font = Font(bold=True)
            self.plan_ativa[cel] = value[0]
            self.plan_ativa[cel].fill = my_fill = PatternFill(patternType='solid', fgColor=value[1])
            self.plan_ativa.column_dimensions[get_column_letter(value[2])].width = len(value[0])+5

    def preencherDados(self):
        for j, i in enumerate(self.__dados['id'].values):
            dados = self.__dados[self.__dados['id'] == i]
            self.inserirDecreto(self.carregarDados(dados, j+3))

    def carregarDados(self, dados, line):
        self.dadosPlan = {
            1 :[dados['N_DECRETO'].values[0], f'A{line}'],
            2: [dados['DATA_ALTERACAO_ORCAMENTARIA'].values[0], f'B{line}'],
            3: [f"{dados['ORGAO_ANULADO'].values[0]} - {dados['NOME_ORGAO_ANULADO'].values[0]}", f'C{line}'],
            4: [f"{dados['ID_PROGRAMA_ANULADO'].values[0]} - {dados['PROGRAMA_ANULADO'].values[0]}", f'D{line}'],
            5: [f"{dados['ID_ACAO_ANULADO'].values[0]} - {dados['ACAO_ANULADO'].values[0]}", f'E{line}'],
            6: [dados['NOME_PRODUTO_ANULADO'].values[0], f'F{line}'],
            7: [dados['VALOR_FISICO_ATUAL_ANULADO'].values[0], f'G{line}'],
            8: [dados['NOVA_META_FISICA_ANULADO'].values[0], f'H{line}'],
            9: [f"{dados['NOME_ORGAO_SUPLEMENTADO'].values[0]} - {dados['NOME_ORGAO_SUPLEMENTADO2'].values[0]}", f'I{line}'],
            10: [f"{dados['ID_PROGRAMA_SUPLEMENTADO'].values[0]} - {dados['PROGRAMA_SUPLEMENTADO'].values[0]}", f'J{line}'],
            11: [f"{dados['ID_ACAO_SUPLEMENTADO'].values[0]} - {dados['ACAO_SUPLEMENTADO'].values[0]}", f'K{line}'],
            12: [dados['NOME_PRODUTO_SUPLEMENTADO'].values[0], f'L{line}'],
            13: [dados['VALOR_FISICO_ATUAL_SUPLEMENTADO'].values[0], f'M{line}'],
            14: [dados['NOVA_META_FISICA_SUPLEMENTADO'].values[0], f'N{line}'],
            15: ["", f'O{line}'],
            16: [dados['EMAIL_INICIAL'].values[0], f'P{line}'],
            17: [dados['VALOR_FINANCEIRO'].values[0], f'Q{line}'],
        }

        try:
            self.dadosPlan[15][0] = dados['DATA_EMAIL_INICIAL'].values[0].strftime('%d/%m/%Y')
        except AttributeError:
            self.dadosPlan[15][0] = dados['DATA_EMAIL_INICIAL'].values[0]

        return self.dadosPlan

    def inserirDecreto(self, dados) -> None:
        for numColumn, value in dados.items():
            self.plan_ativa[value[1]] = value[0]

    def preencherPlanilha(self) -> None:
        self.cabecalho()
        self.preencherDados()
        self.salvar()
